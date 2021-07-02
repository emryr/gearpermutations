# -*- coding: utf-8 -*-
"""
Spyder Editor
"""
import itertools
#from xlwt import Workbook
from openpyxl import load_workbook
import numpy as np
import pandas as pd

#name the workbook z
path = "C:\\Users\\Psych account\\"
wb = load_workbook(path + '20210613_gearpermutation2.xlsx')


#Add a sheet and name it. Allow cell overwrite
permwb = wb.create_sheet('gearpermutations', 0)

#initiating all variables

#Gear names - can be changed to whatever you'd like to label the gears
column = ['compound type','Speed Ratio','Gear 1', 'Gear 2', 'Gear 3', 'Gear 4', 'Gear 5', 'Gear 6', 'Gear 7'] 

values = [2, 4, 6, 8, 10, 16, 24] #array of gears w desired teeth numbers

per = [] #storage of the number of permultations given len(values) choose i
sol = [] #solution array
gearspeed = 1
index = 0

def genbin(n, bs=''):
    if len(bs) == n:
        binom.append(bs)
    else:
        genbin(n, bs + '0')
        genbin(n, bs + '1')

#For loop which stores all the permutations in which you can choose i of the length of the values array
for choose in range(1, len(values)+1):
    binom = []

    #per is a forced list of a map which forces all the tuples in itertools to lists in python - a bit clunky, but functions
    per =  list(map(list, itertools.permutations(values, choose)))
    genbin(choose-1)
    binom = list(map(str, binom))

    print(len(binom), binom)

    #itereates over the number of required replications of possible compound gears 
    for currentperm in range(len(per)):
        #iterates over the length of the permutations given len(values) choose (choose) from above
        for rep in range(len(binom)):               
            sol.append(list(per[currentperm])) #appends per at a given index to the solution array
            sol[-1].insert(0,binom[rep]) #inserts the replication number to track which combo is happening - not elegant 
            gearspeed = 1
   
            #Appends the calculated speed to the speed array, checks if rep = 0
            #checks if the replication is 0, because the calculation for no compunds is different than with compounds
            for index in range(len(per[currentperm])-1):
                #print(binom)
                if binom[rep][index] == '0': 
                    gearspeed = gearspeed*(per[currentperm][index]/per[currentperm][index+1])
                                
            sol[-1].insert(1, gearspeed)


print(len(sol))

df = pd.DataFrame(sol, columns = column)
sort = df.sort_values(by = ['Speed Ratio'])
print(sort)

sort.to_csv("C:/Users/Psych account/Documents/gearpermutations/20210107gearpermutations.csv")                  

